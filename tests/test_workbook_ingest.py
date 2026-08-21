from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from retrieval.ingest.excel.workbook import _main_region_end, build_workbook_chunks


class FakeTokenizer:
    def encode(self, text, **_kwargs):
        return text.split()

    def decode(self, tokens, **_kwargs):
        return " ".join(tokens)


class WorkbookIngestTests(unittest.TestCase):
    def test_main_region_ignores_isolated_cell_after_huge_gap(self):
        self.assertEqual(_main_region_end({1, 2, 3, 1_048_565}), 3)
        self.assertEqual(_main_region_end(set(range(1, 20_001))), 20_000)

    def test_preserves_sheet_rows_and_skips_empty_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Targets"
            sheet.append(["Metric", "Year", "Value"])
            sheet.append(["Pole replacement", 2025, 100])
            sheet.append([None, None, None])
            workbook.save(path)
            workbook.close()

            chunks = build_workbook_chunks(path, FakeTokenizer(), max_tokens=100)

        self.assertEqual(len(chunks), 1)
        self.assertEqual(chunks[0].sheet, "Targets")
        self.assertEqual((chunks[0].row_start, chunks[0].row_end), (2, 2))
        self.assertEqual(chunks[0].header_row_start, 1)
        self.assertEqual(chunks[0].cell_range, "A1:C2")
        self.assertIn("| Metric | Year | Value |", chunks[0].content)
        self.assertIn("Pole replacement", chunks[0].content)

    def test_ignores_formatting_only_cells_at_excel_limits(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "formatted-range.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Targets"
            sheet.append(["Metric", "Value"])
            sheet.append(["Detailed inspection", 100])
            sheet["XFD1048576"].number_format = "0.00"
            workbook.save(path)
            workbook.close()

            chunks = build_workbook_chunks(path, FakeTokenizer(), max_tokens=100)

        self.assertEqual(len(chunks), 1)
        self.assertEqual((chunks[0].row_start, chunks[0].row_end), (2, 2))
        self.assertNotIn("1048576", chunks[0].content)

    def test_repeats_headers_and_preserves_formula_grid(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "formulas.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Costs"
            sheet.append(["Program", "Cost"])
            sheet.append(["A", "=1+2"])
            sheet.append(["B", "=3+4"])
            workbook.save(path)
            workbook.close()

            chunks = build_workbook_chunks(path, FakeTokenizer(), max_tokens=25)

        self.assertGreaterEqual(len(chunks), 2)
        self.assertTrue(all("| Program | Cost |" in chunk.content for chunk in chunks))
        formulas = [formula for chunk in chunks for row in chunk.formula_grid for formula in row]
        self.assertIn("=1+2", formulas)
        self.assertIn("=3+4", formulas)

    def test_uses_generated_headers_when_sheet_has_no_textual_header(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "numbers.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([2024, 10])
            sheet.append([2025, 20])
            workbook.save(path)
            workbook.close()

            chunks = build_workbook_chunks(path, FakeTokenizer(), max_tokens=100)

        self.assertEqual(chunks[0].header_row_start, None)
        self.assertIn("| Column A | Column B |", chunks[0].content)
        self.assertEqual(chunks[0].cell_range, "A1:B2")

    def test_does_not_choose_later_mixed_data_row_as_header(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "mixed.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Attribute", "Sub-attribute", "2025 Value", "2026 Value"])
            sheet.append(["Safety", "Fatality", "$15 million", "$16 million"])
            sheet.append(["Financial", "US Dollar", 1, 1])
            workbook.save(path)
            workbook.close()

            chunks = build_workbook_chunks(path, FakeTokenizer(), max_tokens=100)

        self.assertEqual(chunks[0].header_row_start, 1)
        self.assertIn("| Attribute | Sub-attribute | 2025 Value | 2026 Value |", chunks[0].content)


if __name__ == "__main__":
    unittest.main()
