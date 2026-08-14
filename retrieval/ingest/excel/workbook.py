"""Generic local XLSX ingestion into the shared retrieval index.

The reviewed ``sdge_tableNN`` CSV pipeline remains the authoritative path for
the quarterly-data-report tables.  This module covers additional workbooks in
the local ``files/`` corpus without pretending they follow those contracts.
Rows are preserved as labelled text and their exact sheet/row provenance is
stored in ``structured_data``.
"""

from __future__ import annotations

import argparse
import hashlib
import logging
from xml.etree.ElementTree import iterparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from psycopg2.extras import Json, execute_values
from tqdm import tqdm

from retrieval.contextual_embeddings import (
    CONTEXTUAL_EMBEDDING_RECIPE,
    contextual_embedding_text_for_model,
)
from retrieval.ingest.excel.schema import migrate
from retrieval.setup_db import setup_database
from retrieval.utils import connect_db, embedding_config, get_embedding_model


logger = logging.getLogger(__name__)
EXTRACTOR = "openpyxl-generic-v1"
INGEST_RECIPE = "generic-xlsx-rows-v1"


@dataclass(frozen=True)
class WorkbookChunk:
    sheet: str
    row_start: int
    row_end: int
    content: str
    token_count: int


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return " ".join(str(value).split())


def _row_text(row_number: int, values: Sequence[object]) -> str:
    cells = []
    for index, value in enumerate(values, start=1):
        text = _cell_text(value)
        if text:
            cells.append(f"C{index}={text}")
    return f"Row {row_number}: " + " | ".join(cells)


def _token_count(tokenizer, text: str) -> int:
    return len(tokenizer.encode(text, add_special_tokens=True))


def _nonempty_bounds(workbook, worksheet) -> tuple[int, int]:
    """Return bounds based on values/formulas, ignoring formatting-only cells.

    Regulatory workbooks sometimes apply formatting to Excel's final row or
    column. ``openpyxl`` then reports a million rows or 16,384 columns and its
    read-only iterator materializes every missing cell. Reading the worksheet
    XML directly lets us bound iteration by cells that contain actual data.
    """
    populated_rows: set[int] = set()
    populated_columns: set[int] = set()
    with workbook._archive.open(worksheet._worksheet_path) as source:
        for _, element in iterparse(source, events=("end",)):
            local_name = element.tag.rsplit("}", 1)[-1]
            if local_name == "c":
                has_content = False
                for child in element:
                    child_name = child.tag.rsplit("}", 1)[-1]
                    if child_name == "f":
                        has_content = True
                        break
                    if child_name == "v" and child.text is not None:
                        has_content = True
                        break
                    if child_name == "is" and "".join(child.itertext()).strip():
                        has_content = True
                        break
                coordinate = element.attrib.get("r")
                if has_content and coordinate:
                    row, column = coordinate_to_tuple(coordinate)
                    populated_rows.add(row)
                    populated_columns.add(column)
                element.clear()
            elif local_name == "row":
                element.clear()
    return _main_region_end(populated_rows), _main_region_end(populated_columns)


def _main_region_end(indices: set[int], *, gap_limit: int = 10_000) -> int:
    """Ignore isolated cells separated from the main used region by a huge gap."""
    if not indices:
        return 0
    ordered = sorted(indices)
    end = ordered[0]
    for value in ordered[1:]:
        if value - end > gap_limit:
            break
        end = value
    return end


def _bounded_text_parts(tokenizer, prefix: str, text: str, max_tokens: int) -> list[str]:
    """Split even an unbroken formula/string by model tokens, not whitespace."""
    prefix_tokens = _token_count(tokenizer, prefix)
    allowance = max(1, max_tokens - prefix_tokens - 2)
    token_ids = tokenizer.encode(text, add_special_tokens=False)
    if len(token_ids) <= allowance:
        return [text]
    if not hasattr(tokenizer, "decode"):
        words = text.split()
        return [
            " ".join(words[index : index + allowance])
            for index in range(0, len(words), allowance)
        ]
    return [
        tokenizer.decode(
            token_ids[index : index + allowance], skip_special_tokens=True
        )
        for index in range(0, len(token_ids), allowance)
    ]


def build_workbook_chunks(path: Path, tokenizer, max_tokens: int) -> list[WorkbookChunk]:
    """Read every non-empty worksheet row and pack it into bounded chunks."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    chunks: list[WorkbookChunk] = []
    try:
        for worksheet in workbook.worksheets:
            max_row, max_column = _nonempty_bounds(workbook, worksheet)
            if max_row == 0 or max_column == 0:
                continue
            prefix = f"Workbook: {path.name}\nSheet: {worksheet.title}\n"
            pending: list[tuple[int, str]] = []

            def flush() -> None:
                if not pending:
                    return
                body = "\n".join(text for _, text in pending)
                content = prefix + body
                chunks.append(
                    WorkbookChunk(
                        sheet=worksheet.title,
                        row_start=pending[0][0],
                        row_end=pending[-1][0],
                        content=content,
                        token_count=_token_count(tokenizer, content),
                    )
                )
                pending.clear()

            rows = worksheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
                values_only=True,
            )
            for row_number, row in enumerate(rows, start=1):
                text = _row_text(row_number, row)
                if text.endswith(": "):
                    continue
                candidate = prefix + "\n".join(
                    [*(item[1] for item in pending), text]
                )
                if pending and _token_count(tokenizer, candidate) > max_tokens:
                    flush()
                    candidate = prefix + text
                if _token_count(tokenizer, candidate) > max_tokens:
                    # Preserve a pathological wide row in bounded pieces.  The
                    # sheet and row number are repeated in every piece.
                    flush()
                    for part in _bounded_text_parts(
                        tokenizer, prefix, text, max_tokens
                    ):
                        content = prefix + part
                        chunks.append(
                            WorkbookChunk(
                                worksheet.title,
                                row_number,
                                row_number,
                                content,
                                _token_count(tokenizer, content),
                            )
                        )
                else:
                    pending.append((row_number, text))
            flush()
    finally:
        workbook.close()
    return chunks


def _embed(model, path: Path, chunks: Sequence[WorkbookChunk]) -> tuple[list, list]:
    if not chunks:
        return [], []
    raw = model.encode(
        [chunk.content for chunk in chunks],
        normalize_embeddings=True,
        show_progress_bar=False,
    )
    contextual_inputs = [
        contextual_embedding_text_for_model(
            path.name,
            f"{path.name} > {chunk.sheet} > rows {chunk.row_start}-{chunk.row_end}",
            chunk.content,
            model.tokenizer,
            model.max_seq_length,
        )
        for chunk in chunks
    ]
    contextual = model.encode(
        contextual_inputs,
        normalize_embeddings=True,
        show_progress_bar=False,
    )
    return [value.tolist() for value in raw], [value.tolist() for value in contextual]


def ingest_workbook(path: Path, conn, model, *, force: bool = False) -> str:
    source_hash = _file_hash(path)
    signature = hashlib.sha256(
        f"{INGEST_RECIPE}|{embedding_config()}".encode()
    ).hexdigest()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT id, content_hash, ingest_signature FROM documents WHERE filename = %s",
            (path.name,),
        )
        existing = cursor.fetchone()
    if existing and existing[1:] == (source_hash, signature) and not force:
        logger.info("Skipping unchanged workbook %s", path.name)
        return "unchanged"

    max_tokens = int(embedding_config()["max_tokens_per_chunk"])
    chunks = build_workbook_chunks(path, model.tokenizer, max_tokens)
    raw_embeddings, contextual_embeddings = _embed(model, path, chunks)
    config = embedding_config()
    with conn.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO documents (
                filename, page_count, content_hash, ingest_signature, chunk_counts
            ) VALUES (%s, 0, %s, %s, %s)
            ON CONFLICT (filename) DO UPDATE
            SET content_hash = EXCLUDED.content_hash,
                ingest_signature = EXCLUDED.ingest_signature,
                chunk_counts = EXCLUDED.chunk_counts,
                ingested_at = now()
            RETURNING id
            """,
            (path.name, source_hash, signature, Json({"excel_card": len(chunks)})),
        )
        document_id = cursor.fetchone()[0]
        cursor.execute("DELETE FROM chunks WHERE document_id = %s", (document_id,))
        rows = []
        for index, (chunk, raw, contextual) in enumerate(
            zip(chunks, raw_embeddings, contextual_embeddings, strict=True)
        ):
            breadcrumb = (
                f"{path.name} > {chunk.sheet} > rows "
                f"{chunk.row_start}-{chunk.row_end}"
            )
            rows.append(
                (
                    document_id,
                    chunk.sheet,
                    breadcrumb,
                    None,
                    0,
                    0,
                    index,
                    "excel_card",
                    chunk.content,
                    None,
                    f"{chunk.sheet}, rows {chunk.row_start}-{chunk.row_end}",
                    Json(
                        {
                            "kind": "workbook_rows",
                            "source_file": path.name,
                            "sheet": chunk.sheet,
                            "row_start": chunk.row_start,
                            "row_end": chunk.row_end,
                        }
                    ),
                    None,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    chunk.token_count,
                    hashlib.sha256(chunk.content.encode()).hexdigest(),
                    config.get("provider", "sentence_transformers"),
                    config["name"],
                    list(raw),
                    config["name"],
                    CONTEXTUAL_EMBEDDING_RECIPE,
                    list(contextual),
                    EXTRACTOR,
                )
            )
        if rows:
            execute_values(
                cursor,
                """
                INSERT INTO chunks (
                    document_id, sub_document, breadcrumb, section_number,
                    page_start, page_end, chunk_index, content_type,
                    content, retrieval_hint, caption, structured_data,
                    object_key, media_type, token_count, content_hash,
                    embedding_provider, embedding_model, embedding,
                    contextual_embedding_model, contextual_embedding_recipe,
                    contextual_embedding, extractor
                ) VALUES %s
                """,
                rows,
            )
    conn.commit()
    logger.info("Ingested %s: %d workbook chunks", path.name, len(chunks))
    return "loaded"


def ingest_directory(
    directory: Path,
    conn,
    model,
    *,
    force: bool = False,
    exclude_names: set[str] | None = None,
) -> dict[str, str]:
    excluded = exclude_names or set()
    paths = sorted(
        path
        for path in [*directory.rglob("*.xlsx"), *directory.rglob("*.xlsm")]
        if path.name not in excluded
    )
    results: dict[str, str] = {}
    for path in tqdm(paths, desc="Ingesting workbooks", unit="file"):
        try:
            results[path.name] = ingest_workbook(path, conn, model, force=force)
        except Exception:
            conn.rollback()
            logger.exception("Failed to ingest workbook %s", path.name)
            results[path.name] = "failed"
    return results


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_dir", type=Path, nargs="?", default=Path("files"))
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    setup_database()
    conn = connect_db()
    try:
        migrate(conn)
        results = ingest_directory(
            args.input_dir, conn, get_embedding_model(), force=args.force
        )
    finally:
        conn.close()
    if not results:
        raise SystemExit(f"No XLSX/XLSM files matched in {args.input_dir}")
    failed = [name for name, status in results.items() if status == "failed"]
    if failed:
        raise SystemExit(f"Workbook ingestion failed: {', '.join(failed)}")


if __name__ == "__main__":
    main()
